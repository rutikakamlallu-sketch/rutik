import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
import os

# Read the CSV file
csv_file = '3-3+forecast_table_raw.csv'
df = pd.read_csv(csv_file)

# Create Excel workbook
excel_file = 'Sales_Forecast_Dashboard.xlsx'
with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Raw Data', index=False)
    
    # Create Summary sheet
    summary_data = {
        'Metric': ['Revenue', 'COGS', 'Opex', 'EBITDA', 'Cash_End'],
        'Q1_Actual': [120000, 72000, 28000, 20000, 12000],
        'Q2_Plan': [130000, 76000, 29500, 24500, 12500],
        'Q2_Forecast': [128000, 77000, 29000, 22000, 11400],
        'Plan_vs_Forecast': [2000, -1000, 500, 2500, 1100],
        'Variance_%': [1.54, -1.32, 1.69, 10.20, 8.80],
        'P80_Low': [124000, 75000, 28500, 19500, 10200],
        'P80_High': [132000, 79000, 29500, 23500, 12800]
    }
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    # Create Drivers sheet
    drivers_data = {
        'Metric': ['Revenue', 'COGS', 'Opex', 'EBITDA'],
        'Driver_1': ['Price: +2000', 'Materials: +1500', 'Hiring: +700', 'Cumulative Effect'],
        'Driver_2': ['Volume: -2500', 'Efficiency: -500', 'Marketing: -200', ''],
        'Driver_3': ['FX: -1500', 'FX: +1000', '', ''],
        'Impact_Summary': ['Net: -2000', 'Net: +2000', 'Net: +500', 'Net: +500']
    }
    drivers_df = pd.DataFrame(drivers_data)
    drivers_df.to_excel(writer, sheet_name='Drivers', index=False)

# Load workbook for formatting
wb = openpyxl.load_workbook(excel_file)

# ===== FORMAT SUMMARY SHEET =====
ws_summary = wb['Summary']
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)
center_alignment = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Format headers
for cell in ws_summary[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border

# Format data rows with currency and percentage
for row in range(2, ws_summary.max_row + 1):
    for col in range(1, ws_summary.max_column + 1):
        cell = ws_summary.cell(row, col)
        cell.border = thin_border
        cell.alignment = center_alignment
        
        if col > 1 and col <= 6:  # Currency columns
            cell.number_format = '#,##0'
        elif col == 6:  # Variance % column
            cell.number_format = '0.00%'

# Adjust column widths
ws_summary.column_dimensions['A'].width = 15
for col in range(2, ws_summary.max_column + 1):
    ws_summary.column_dimensions[get_column_letter(col)].width = 14

# ===== FORMAT DRIVERS SHEET =====
ws_drivers = wb['Drivers']
for cell in ws_drivers[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border

for row in range(2, ws_drivers.max_row + 1):
    for col in range(1, ws_drivers.max_column + 1):
        cell = ws_drivers.cell(row, col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

ws_drivers.column_dimensions['A'].width = 12
for col in range(2, ws_drivers.max_column + 1):
    ws_drivers.column_dimensions[get_column_letter(col)].width = 20
ws_drivers.row_dimensions[1].height = 30

# ===== CREATE CHARTS =====
# Revenue Comparison Chart
ws_chart = wb['Summary']
bar_chart = BarChart()
bar_chart.type = 'col'
bar_chart.style = 10
bar_chart.title = 'Revenue: Actual vs Plan vs Forecast'
bar_chart.y_axis.title = 'Amount ($)'
bar_chart.x_axis.title = 'Quarters'

data = Reference(ws_chart, min_col=2, min_row=1, max_row=2, max_col=4)
categories = Reference(ws_chart, min_col=1, min_row=2, max_row=2)
bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(categories)
ws_chart.add_chart(bar_chart, 'E2')

# EBITDA Trend Chart
line_chart = LineChart()
line_chart.title = 'EBITDA Trend & Forecast'
line_chart.y_axis.title = 'EBITDA ($)'
line_chart.style = 11

ebitda_data = Reference(ws_chart, min_col=2, min_row=4, max_row=4, max_col=4)
ebitda_cats = Reference(ws_chart, min_col=1, min_row=4, max_row=4)
line_chart.add_data(ebitda_data, titles_from_data=True)
line_chart.set_categories(ebitda_cats)
ws_chart.add_chart(line_chart, 'E16')

# Save workbook
wb.save(excel_file)
print(f"✅ Excel Dashboard created successfully: {excel_file}")
print(f"📊 Sheets created: Raw Data, Summary, Drivers")
print(f"📈 Charts added: Revenue Comparison, EBITDA Trend")
