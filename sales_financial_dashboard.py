import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import os

# Read the CSV files
accounts_df = pd.read_csv('Accounts.csv')
forecast_df = pd.read_csv('3-3+forecast_table_raw.csv')

# Create Excel workbook
excel_file = 'Sales_Financial_Dashboard.xlsx'
with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    # Sheet 1: Account Structure
    accounts_df.to_excel(writer, sheet_name='Account Structure', index=False)
    
    # Sheet 2: Forecast Data
    forecast_df.to_excel(writer, sheet_name='Forecast Data', index=False)
    
    # Sheet 3: P&L Statement (Sample Data)
    pl_data = {
        'Account': [
            'Gross Revenue',
            'Cost of Goods Sold',
            'Gross Profit',
            'Research and Development',
            'Sales and Marketing',
            'General and Administrative',
            'Salaries and Wages',
            'Interest',
            'Rent and Utilities',
            'Other',
            'Total Operating Expenses',
            'Operating Income (EBIT)',
            'Taxes',
            'Income Before Taxes',
            'Net Income'
        ],
        'Q1_Actual': [120000, -72000, 48000, -5000, -8000, -4000, -6000, -2000, -2000, -1000, -28000, 20000, -3000, 17000, 14000],
        'Q2_Forecast': [128000, -77000, 51000, -5500, -8500, -4200, -6500, -2200, -2100, -1100, -29500, 21500, -3225, 18275, 15050],
        'Q2_Plan': [130000, -76000, 54000, -5200, -8200, -4100, -6300, -2100, -2000, -900, -29500, 24500, -3675, 20825, 17150],
        'Variance_$': [-2000, -1000, -3000, -300, -300, -100, -200, -100, -100, -200, 0, -3000, 450, -2550, -2100],
        'Variance_%': [-1.54, 1.32, -5.56, 5.77, 3.66, 2.44, 3.17, 4.76, 5.00, 22.22, 0.00, -12.24, -12.24, -12.24, -12.24]
    }
    pl_df = pd.DataFrame(pl_data)
    pl_df.to_excel(writer, sheet_name='P&L Summary', index=False)
    
    # Sheet 4: Expense Breakdown
    expense_data = {
        'Expense Category': [
            'Research and Development',
            'Sales and Marketing',
            'General and Administrative',
            'Salaries and Wages',
            'Interest',
            'Rent and Utilities',
            'Other'
        ],
        'Q1_Actual': [5000, 8000, 4000, 6000, 2000, 2000, 1000],
        'Q2_Forecast': [5500, 8500, 4200, 6500, 2200, 2100, 1100],
        'Q2_Plan': [5200, 8200, 4100, 6300, 2100, 2000, 900],
        'Budget_%': [17.6, 28.0, 14.2, 21.5, 7.4, 7.1, 3.7]
    }
    expense_df = pd.DataFrame(expense_data)
    expense_df.to_excel(writer, sheet_name='Expense Breakdown', index=False)
    
    # Sheet 5: Dashboard Summary
    dashboard_data = {
        'KPI': [
            'Gross Revenue',
            'Gross Profit',
            'Gross Profit Margin %',
            'Operating Income (EBIT)',
            'EBIT Margin %',
            'Net Income',
            'Net Margin %',
            'Total Operating Expenses'
        ],
        'Q1_Actual': [120000, 48000, 40.0, 20000, 16.7, 14000, 11.7, 28000],
        'Q2_Forecast': [128000, 51000, 39.8, 21500, 16.8, 15050, 11.8, 29500],
        'Q2_Plan': [130000, 54000, 41.5, 24500, 18.8, 17150, 13.2, 29500],
        'Forecast_vs_Plan': [-2000, -3000, -1.7, -3000, -2.0, -2100, -1.4, 0],
        'Status': ['⚠️ Below', '⚠️ Below', '⚠️ Below', '⚠️ Below', '⚠️ Below', '⚠️ Below', '⚠️ Below', '✅ On Track']
    }
    dashboard_df = pd.DataFrame(dashboard_data)
    dashboard_df.to_excel(writer, sheet_name='Executive Summary', index=False)

# Load workbook for formatting
wb = openpyxl.load_workbook(excel_file)

# Define styles
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
subheader_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
subheader_font = Font(bold=True, color='FFFFFF', size=10)
accent_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
warning_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
success_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
center_alignment = Alignment(horizontal='center', vertical='center')
left_alignment = Alignment(horizontal='left', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ===== FORMAT EXECUTIVE SUMMARY =====
ws_exec = wb['Executive Summary']
for cell in ws_exec[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border

for row in range(2, ws_exec.max_row + 1):
    for col in range(1, ws_exec.max_column + 1):
        cell = ws_exec.cell(row, col)
        cell.border = thin_border
        
        if col == 1:
            cell.alignment = left_alignment
            cell.font = Font(bold=True)
        else:
            cell.alignment = center_alignment
            if col <= 4 and col > 1:
                cell.number_format = '#,##0'
            elif col == 5 or col == 6:
                cell.number_format = '0.0'
            elif col == 7:
                cell.alignment = center_alignment

ws_exec.column_dimensions['A'].width = 25
for col in range(2, ws_exec.max_column + 1):
    ws_exec.column_dimensions[get_column_letter(col)].width = 14

# ===== FORMAT P&L SUMMARY =====
ws_pl = wb['P&L Summary']
for cell in ws_pl[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border

for row in range(2, ws_pl.max_row + 1):
    for col in range(1, ws_pl.max_column + 1):
        cell = ws_pl.cell(row, col)
        cell.border = thin_border
        
        if col == 1:
            cell.alignment = left_alignment
            cell.font = Font(bold=True)
            # Highlight key totals
            if cell.value in ['Gross Profit', 'Total Operating Expenses', 'Operating Income (EBIT)', 'Net Income']:
                cell.fill = accent_fill
        else:
            cell.alignment = center_alignment
            if col <= 4:
                cell.number_format = '#,##0'
            elif col >= 5:
                cell.number_format = '0.0'

ws_pl.column_dimensions['A'].width = 30
for col in range(2, ws_pl.max_column + 1):
    ws_pl.column_dimensions[get_column_letter(col)].width = 13

# ===== FORMAT EXPENSE BREAKDOWN =====
ws_expense = wb['Expense Breakdown']
for cell in ws_expense[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border

for row in range(2, ws_expense.max_row + 1):
    for col in range(1, ws_expense.max_column + 1):
        cell = ws_expense.cell(row, col)
        cell.border = thin_border
        
        if col == 1:
            cell.alignment = left_alignment
        else:
            cell.alignment = center_alignment
            if col <= 4:
                cell.number_format = '#,##0'
            else:
                cell.number_format = '0.0%'

ws_expense.column_dimensions['A'].width = 28
for col in range(2, ws_expense.max_column + 1):
    ws_expense.column_dimensions[get_column_letter(col)].width = 13

# ===== CREATE CHARTS =====

# Chart 1: Revenue & Profitability Trend
ws_charts = wb['P&L Summary']
bar_chart1 = BarChart()
bar_chart1.type = 'col'
bar_chart1.style = 10
bar_chart1.title = 'Revenue & Profitability Comparison'
bar_chart1.y_axis.title = 'Amount ($)'

data1 = Reference(ws_charts, min_col=2, min_row=1, max_row=2, max_col=4)
bar_chart1.add_data(data1, titles_from_data=True)
bar_chart1.set_categories(Reference(ws_charts, min_col=1, min_row=2))
ws_charts.add_chart(bar_chart1, 'G2')

# Chart 2: Expense Breakdown Pie Chart
ws_exp = wb['Expense Breakdown']
pie_chart = PieChart()
pie_chart.title = 'Q2 Forecast Expense Distribution'

labels = Reference(ws_exp, min_col=1, min_row=2, max_row=8)
data2 = Reference(ws_exp, min_col=3, min_row=1, max_row=8)
pie_chart.add_data(data2, titles_from_data=True)
pie_chart.set_categories(labels)
ws_exp.add_chart(pie_chart, 'F2')

# Chart 3: Margin Analysis Line Chart
line_chart = LineChart()
line_chart.title = 'Margin Analysis: Forecast vs Plan'
line_chart.y_axis.title = 'Margin %'
line_chart.style = 11

ws_margin = wb['Executive Summary']
margin_data = Reference(ws_margin, min_col=3, min_row=1, max_row=3, max_col=4)
line_chart.add_data(margin_data, titles_from_data=True)
ws_margin.add_chart(line_chart, 'G2')

# Save workbook
wb.save(excel_file)
print(f"✅ Sales & Financial Dashboard created: {excel_file}")
print(f"📊 Sheets created: Account Structure, Forecast Data, P&L Summary, Expense Breakdown, Executive Summary")
print(f"📈 Charts added: Revenue & Profitability, Expense Distribution, Margin Analysis")
print(f"\n🎯 Key Features:")
print(f"   • 6-level account hierarchy tracking")
print(f"   • P&L statement with variance analysis")
print(f"   • Expense breakdown by category")
print(f"   • Executive summary with KPIs")
print(f"   • Professional formatting & charts")
